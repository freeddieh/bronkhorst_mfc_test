import openpyxl
import datetime
import pandas as pd
import tkinter as tk
from airpy import *
from tkinter.filedialog import askopenfilename


def before_after(timestamp) -> bool:
    """
    
    Tests whether or not an input timestamp lies before or after the current time.

    :return: True if timestamp lies after the current time, False if it lies before timestamp
    
    """
    if timestamp >= datetime.datetime.now():
        return True
    else:
        return False


def find_program_index(file: str, program: str) -> list[int]:
    workbook = openpyxl.load_workbook(filename=file, data_only=True)

    # Select the leftmost sheet
    sheet = workbook.worksheets[0]  # The first sheet in the workbook

    # Create a DataFrame from the sheet data
    data = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx < 1:
            header = row
        elif 1 <= idx <= 51:
            data.append(row)

    # Create a DataFrame with the data
    df = pd.DataFrame(data, columns=header)

    # Identify colored cells and modify the DataFrame in place
    for row in sheet.iter_rows(max_row=51):
        for cell in row:
            cell_color = cell.fill.start_color.index
            if str(cell_color) not in ['00000000', '0', '43']:  # Check if the cell has a fill color
                # Modify the corresponding DataFrame cell
                df.iat[cell.row-2, cell.column-1] = 'Colored'
            else:
                df.iat[cell.row-2, cell.column-1] = None
    
    if 'Skift' not in program:
        program_steps = df[df[program]=='Colored'].index
    elif 'Skift' in program:
        program_steps = df[df['Inds v skift af ref']=='Colored'].index

    return program_steps


def find_mfc_setpoints() -> tuple[list]:
    # Names of columns containing the setpoint values in %
    setp1_name = 'Fd %'
    setp2_name = 'Fso2'

    # Names of columns containing the setpoint values in ml/min 
    
    #setv1_name = 'ml/min'
    #setv2_name = 'ml/min'    
    # As the column names are currently identical in the Excel sheet 
    # this is less than ideal for this use, as they cannot be distingushed from eachother
    
    # Find instruction file and select input type:
    root_file = tk.Tk()
    root_file.withdraw()
    root_file.wm_attributes('-topmost', 1)
    ark = askopenfilename(parent=root_file, title='Vælg fil med instrukserne')
    root_file.destroy()

    # Create the main window
    options = ['Liniaritet', 'Span', 'Reference Gas Skift']
    
    root_prog = tk.Tk()
    root_prog.geometry('500x300')
    root_prog.wm_attributes('-topmost', 1)
    root_prog.title('Blandingsprogram')

    # Instantiate the DropdownApp class
    ddm = DropdownMenu(root_prog, options)

    # Start the Tkinter event loop
    root_prog.mainloop()

    program = ddm.get_selected_value()

    # Construct DataFrame with the loaded sheet an slice for readability
    df = pd.read_excel(ark)
    for idx, item in enumerate(df.iloc[:, 0]):
        try:
            int(item)
        except ValueError:
            slicer = idx
            break
    df = df.loc[:slicer-1]

    indexes = find_program_index(ark, program)
    
    #setv1 = df[setv1_name].iloc[indexes].values
    #setv2 = df[setv2_name].iloc[indexes].values
    
    setp1 = df[setp1_name].iloc[indexes].values
    setp2 = df[setp2_name].iloc[indexes].values

    return setp1, setp2 #, setv1, setv2


def main() -> None:
    """
    
    Defines the main function to controll the Bronkhorst MFC's using the worksheet
    
    """
    
    bh_ports = list(find_bronkhorst_ports().values())

    bronkhorst_small = BronkhorstMFC(bh_ports[0])
    bronkhorst_large = BronkhorstMFC(bh_ports[1])
    
    for dilution, span in zip(*find_mfc_setpoints):
        bronkhorst_large.write_bronkhorst(dilution)
        bronkhorst_small.write_bronkhorst(span)


if __name__ == '__main__':

    for values in zip(*find_mfc_setpoints()):
        print(values)