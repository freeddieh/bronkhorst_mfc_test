import sys
import time
import openpyxl
import datetime
import pandas as pd
import tkinter as tk
from airpy import *
from tkinter.filedialog import askopenfilename


def before_after(timestamp) -> bool:
    """
    
    Tests whether or not an input timestamp lies before or after the current time.

    :return: True if timestamp lies after the current time, False if it lies before timestamp
    
    """
    if timestamp >= datetime.datetime.now():
        return True
    else:
        return False


def find_program_index(file: str, program: str) -> tuple[list]:
    """
    
    Finds the indices for setpoints for MFCs in a given program

    :param file: The file with the setpoints for the program
    :param program: The program to find the setpoint indices for 
    
    """
    workbook = openpyxl.load_workbook(filename=file, data_only=True)

    # Select the leftmost sheet
    sheet = workbook.worksheets[0]  # The first sheet in the workbook

    # Create a DataFrame from the sheet data
    data = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx < 1:
            header = row
        elif 1 <= idx <= 51:
            data.append(row)

    # Create a DataFrame with the data
    df = pd.DataFrame(data, columns=header)

    # Identify colored cells and modify the DataFrame in place
    for row in sheet.iter_rows(max_row=51):
        for cell in row:
            cell_color = cell.fill.start_color.index
            if str(cell_color) not in ['00000000', '0', '43']:  # Check if the cell has a fill color
                # Modify the corresponding DataFrame cell
                df.iat[cell.row-2, cell.column-1] = 'Colored'
            else:
                df.iat[cell.row-2, cell.column-1] = None
    
    if 'Skift' not in program:
        program_steps = df[df[program]=='Colored'].index
    elif 'Skift' in program:
        program_steps = df[df['Inds v skift af ref']=='Colored'].index

    return program_steps


def find_mfc_setpoints() -> tuple[str, list[int], list[int]]:
    """
    
    Lets a user select an Excel sheet with instructions for a set of
    calibration programs and lets the user select what specific program
    to run.
    
    :return: A tuple of the chosen program and the setpoints for
             the MFCs in that program
    
    """

    # Names of columns containing the setpoint values in %
    setp1_name = 'Fd %'
    setp2_name = 'Fso2'

    # Names of columns containing the setpoint values in ml/min 
    #setv1_name = 'ml/min'
    #setv2_name = 'ml/min'    
    # As the column names are currently identical in the Excel sheet 
    # this is less than ideal for this use, as they cannot be distingushed from eachother
    
    # Find instruction file and select input type:
    root_file = tk.Tk()
    root_file.withdraw()
    root_file.wm_attributes('-topmost', 1)
    ark = askopenfilename(parent=root_file, title='Vælg fil med instrukserne')
    root_file.destroy()

    # Create the main window
    options = ['Nulstilling', 'Liniaritet', 'Span', 'Reference Gas Skift', 'Afslutning']
    
    root_prog = tk.Tk()
    root_prog.geometry('500x300')
    root_prog.wm_attributes('-topmost', 1)
    root_prog.title('Vælg Blandingsprogram')

    # Instantiate the DropdownApp class
    ddm = DropdownMenu(root_prog, options)

    # Start the Tkinter event loop
    root_prog.mainloop()

    program = ddm.get_selected_value()

    # Construct DataFrame with the loaded sheet an slice for readability
    df = pd.read_excel(ark)
    for idx, item in enumerate(df.iloc[:, 0]):
        try:
            int(item)
        except ValueError:
            slicer = idx
            break
    df = df.loc[:slicer-1]

    if program not in ['Nulstilling', 'Afslutning']:
        indexes = find_program_index(ark, program)
        setp1 = df[setp1_name].iloc[indexes].values
        setp2 = df[setp2_name].iloc[indexes].values
        #setv1 = df[setv1_name].iloc[indexes].values
        #setv2 = df[setv2_name].iloc[indexes].values
    elif program == 'Nulstilling':
        setp1 = [0]
        setp2 = [0]
    elif program == 'Afslutning':
        sys.exit()

    return program, setp1, setp2 #, setv1, setv2


def pct_mln_conversion(mfc_max_flow: float, pct_flow: float) -> float:
    """
    
    Converts a percentage flow from a MFC into a specific flow in mLn/min or Ln/min.

    :param mfc_max_flow: The maximum flow of a given MFC to convert output for.
    :param pct_flow: The percentage setpoint for a MFC to be converted to an specific flow

    :return: Specific flow for the MFC in appropriate unit.
    
    """
    return ((pct_flow / 100) * mfc_max_flow)


def main_controller(bronkhorsts: list[BronkhorstMFC], sleep_time: int = 3600) -> None:
    """
    
    Defines the main function to controll the Bronkhorst MFC's using the worksheet.

    :param bronkhorsts: List of Bronkhorst MFC objects to be controlled.
    :param sleep_time: Time the program is supposed to sleep between dilution steps
                       to achieve a stable concentration.
    
    """
    if len(bronkhorsts) < 1 or len(bronkhorsts) > 2:
        raise KeyError('Uncompatible number of Bronkhorst MFCs connected. Program can only handle 2 MFCs (span + dilution).')
    bronkhorst_small = bronkhorsts[0]
    bronkhorst_large = bronkhorsts[1]
    program, set_pt1, set_pt2 = find_mfc_setpoints()
    set_pts = (set_pt1, set_pt2)
    
    # 206 is the DDE number for setting the specific flow of a Bronkhorst MFC
    for dilution, span in zip(*set_pts):
        dilution_flow = pct_mln_conversion(bronkhorst_large.max_flow, dilution)
        span_flow = pct_mln_conversion(bronkhorst_small.max_flow, span)
        bronkhorst_large.write_bronkhorst(206, dilution_flow)
        bronkhorst_small.write_bronkhorst(206, span_flow)
        if program in ['Nulstilling', 'Afslutning']:
            time.sleep(1)
        else:
            time.sleep(sleep_time)
    
    bronkhorst_large.write_bronkhorst(206, 0)
    bronkhorst_small.write_bronkhorst(206, 0)
    print('Det valgte program er nu fuldendt.')


if __name__ == '__main__':
    main_controller()