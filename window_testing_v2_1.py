import os
import sys
import csv
import glob
import time
import openpyxl
import datetime
import numpy as np
import pandas as pd
import tkinter as tk
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from tkinter import ttk
from tkinter import messagebox
from tkcalendar import Calendar
from bronkhorst_mfc_test.airpy import *
from bronkhorst_mfc_test.mfc_logger_v1 import *
from bronkhorst_mfc_test.mfc_controller_v1 import *
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter.filedialog import askopenfilename


class ProgrammeSelector:
    def __init__(self):
        cal_types = {'SO2': ['Fd %', 'Fso2', 'SO2 ppb'],
                     'NOx': ['Fd %', 'Fno', 'No PPB', 'Nox PPB']}
        # File selection window
        root_file = tk.Tk()
        root_file.withdraw()
        root_file.wm_attributes('-topmost', 1)
        self.ark = askopenfilename(parent=root_file, title='Vælg fil med instrukserne')
        root_file.destroy()

        # Load and slice DataFrame
        try:
            df = pd.read_excel(self.ark)
            for idx, item in enumerate(df.iloc[:, 0]):
                try:
                    int(item)
                except ValueError:
                    slicer = idx
                    break
            df = df.loc[:slicer-1]
        except Exception as e:
            print(f'Error reading Excel file: {e}')
            return
        self.options = df.columns[-3:]

        # Sequential GUI flow
        self.program, self.cal_type, self.clean_program = self._select_programme()
        self.program_steps = self._find_program_index()
        self.selected_starttime, self.time_per_step = self._select_time()
        self.set_point_names = cal_types[self.cal_type]
        self.save_name = f'{self.cal_type}_{self.clean_program}_{self.selected_starttime.strftime("%Y_%m_%d_%H_%M")}'
    

    def _select_programme(self):
        selected_programme = []
        selected_species = []

        def _confirm_selection():
            selected_prog = programme_var.get()
            selected_spec = species_var.get()
            selected_programme.append(selected_prog)
            selected_species.append(selected_spec)
            if selected_prog and selected_spec:
                root_select.destroy()
            else:
                messagebox.showwarning('Ugyldigt valg!', 'Vælg venligst både et program og en type og tryk bekræft.')

        root_select = tk.Tk()
        root_select.wm_attributes('-topmost', 1)
        root_select.title('Vælg Program og Type')
        root_select.geometry('500x300')

        # Main frame with two columns
        main_frame = ttk.Frame(root_select)
        main_frame.pack(expand=True, fill='both', padx=10, pady=10)

        # Programme selection
        programme_frame = ttk.LabelFrame(main_frame, text='Vælg Program')
        programme_frame.pack(side='right', expand=True, fill='both', padx=10)
        programme_var = tk.StringVar()

        for option in self.options:
            if 'skift' in option.lower():
                option_text = 'Reference Gas Skift'
            else:
                option_text = option
            ttk.Radiobutton(programme_frame, text=option_text, variable=programme_var, value=option).pack(anchor='w', pady=5)

        # Species selection
        species_frame = ttk.LabelFrame(main_frame, text='Vælg Type')
        species_frame.pack(side='left', expand=True, fill='both', padx=10)
        species_var = tk.StringVar()
        ttk.Radiobutton(species_frame, text='SO₂', variable=species_var, value='SO2').pack(anchor='w', pady=5)
        ttk.Radiobutton(species_frame, text='NOx', variable=species_var, value='NOx').pack(anchor='w', pady=5)

        # Confirm button
        ttk.Button(root_select, text='Bekræft', command=_confirm_selection).pack(pady=10)

        root_select.mainloop()
        return selected_programme[0], selected_species[0], option_text.replace(' ', '_').lower()


    def _find_program_index(self):
        workbook = openpyxl.load_workbook(filename=self.ark, data_only=True)
        sheet = workbook.worksheets[0]
        data = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx < 1:
                header = row
            elif 1 <= idx <= 51:
                data.append(row)

        df = pd.DataFrame(data, columns=header)
        for row in sheet.iter_rows(max_row=51):
            for cell in row:
                cell_color = cell.fill.start_color.index
                if str(cell_color) not in ['00000000', '0', '43']:
                    df.iat[cell.row-2, cell.column-1] = 'Colored'
                else:
                    df.iat[cell.row-2, cell.column-1] = None

        if 'Skift' not in self.program:
            return df[df[self.program] == 'Colored'].index
        else:
            return df[df['Inds v skift af ref'] == 'Colored'].index


    def _select_time(self):
        selected_datetime = []

        def _confirm_selection():
            if start_mode.get() == 'later':
                selected_date = cal.selection_get()
                selected_time = datetime.time(hour=hour_var.get(), minute=minute_var.get())
                selected_datetime.append(datetime.datetime.combine(selected_date, selected_time))
            else:
                selected_datetime.append(datetime.datetime.now())
            #print(f'Selected Start Time: {selected_datetime}')
            root_time.destroy()

        def _update_ui():
            is_later = start_mode.get() == 'later'
            cal.configure(state='normal' if is_later else 'disabled')
            hour_spinbox.configure(state='normal' if is_later else 'disabled')
            minute_spinbox.configure(state='normal' if is_later else 'disabled')

        root_time = tk.Tk()
        root_time.title('Vælg starttidspunkt for programmet')
        root_time.geometry('450x600')

        # Start mode selection
        start_mode = tk.StringVar(value='now')
        mode_frame = ttk.LabelFrame(root_time, text='Vælg Start')
        mode_frame.pack(padx=10, pady=10, fill='x')

        ttk.Radiobutton(mode_frame, text='Start Nu', variable=start_mode, value='now', command=_update_ui).pack(anchor='w', padx=10)
        ttk.Radiobutton(mode_frame, text='Start Senere', variable=start_mode, value='later', command=_update_ui).pack(anchor='w', padx=10)

        # Calendar
        cal_frame = ttk.LabelFrame(root_time, text='Vælg Dato')
        cal_frame.pack(padx=10, pady=10, fill='both', expand=True)
        cal = Calendar(cal_frame, selectmode='day', date_pattern='yyyy-mm-dd')
        cal.pack(padx=10, pady=10)

        # Time selection
        hour_var = tk.IntVar(value=datetime.datetime.now().hour)
        minute_var = tk.IntVar(value=datetime.datetime.now().minute)

        time_frame = ttk.LabelFrame(root_time, text='Vælg Tid')
        time_frame.pack(padx=10, pady=10, fill='x')
        time_inner_frame = ttk.Frame(time_frame)
        time_inner_frame.pack(pady=10)

        hour_spinbox = ttk.Spinbox(time_inner_frame, from_=0, to=23, textvariable=hour_var, width=5, justify='center')
        hour_spinbox.pack(side='left')
        ttk.Label(time_inner_frame, text=':', font=('Segoe UI', 12)).pack(side='left', padx=5)
        minute_spinbox = ttk.Spinbox(time_inner_frame, from_=0, to=59, textvariable=minute_var, width=5, justify='center')
        minute_spinbox.pack(side='left')

        # Time at each step
        step_time_var = tk.IntVar(value=60)
        step_frame = ttk.LabelFrame(root_time, text='Tid pr. trin (minuter)')
        step_frame.pack(padx=10, pady=10, fill='x')
        ttk.Spinbox(step_frame, from_=1, to=120, textvariable=step_time_var, width=5).pack(pady=5)

        # Expected finish time
        expected_label = ttk.Label(root_time, text='Forventet færdig: --:--')
        expected_label.pack(pady=10)

        def _update_expected_time(*args):
            if start_mode.get() == 'later':
                try:
                    start_dt = datetime.datetime.combine(cal.selection_get(), datetime.time(hour=hour_var.get(), minute=minute_var.get()))
                except Exception:
                    return
            else:
                start_dt = datetime.datetime.now()

            steps = len(getattr(self, 'program_steps', []))
            duration = datetime.timedelta(minutes=step_time_var.get() * steps)
            expected_dt = start_dt + duration
            expected_label.config(text=f'Forventet færdig: {expected_dt.strftime('%Y-%m-%d %H:%M')}')

        # Bind updates
        step_time_var.trace_add('write', _update_expected_time)
        hour_var.trace_add('write', _update_expected_time)
        minute_var.trace_add('write', _update_expected_time)
        start_mode.trace_add('write', _update_expected_time)
        cal.bind('<<CalendarSelected>>', lambda e: _update_expected_time())

        # Confirm button
        ttk.Button(root_time, text='Bekræft', command=_confirm_selection).pack(pady=10)

        _update_ui()
        _update_expected_time()

        root_time.mainloop()
        time_per_step = step_time_var.get()
        return selected_datetime[0], time_per_step


def main_controller(bronkhorsts: list[BronkhorstMFC], 
                    programme: ProgrammeSelector, 
                    end_setpoint_pct: int) -> None:
    '''
    Defines the main function to controll the Bronkhorst MFC's using the worksheet.

    :param bronkhorsts: List of Bronkhorst MFC objects to be controlled.
    :param sleep_time: Time the program is supposed to sleep between dilution steps
                       to achieve a stable concentration.
    '''

    if len(bronkhorsts) != 2:
        raise KeyError('Uncompatible number of Bronkhorst MFCs connected. Program can only handle 2 MFCs (span + dilution).')
    
    def normalize_flow(mfc: BronkhorstMFC):
        # Convert everything to ln/min for comparison
        if mfc.readout_unit == 'mln/min':
            return mfc.max_flow / 1000  # Convert to ln/min
        return mfc.max_flow
    
    # Starting time and flows
    start_time = programme.selected_starttime
    time_list = []
    flow_list_small = []
    flow_list_large = []

    # Sort the list by normalized flow
    mfcs_sorted = sorted(bronkhorsts, key=normalize_flow)
    bronkhorst_small = mfcs_sorted[0]
    bronkhorst_large = mfcs_sorted[1]
    set_pt1, set_pt2, ppb_conc = find_setpoints(programme)
    set_pts = (set_pt1, set_pt2, ppb_conc)
    step_time = programme.time_per_step*60
    csv_header = ['Datetime', 
                  f'Bronkhorst {bronkhorst_small.max_flow}SCCM [mln/min]', 
                  f'Bronkhorst {bronkhorst_large.max_flow}SLM [ln/min]']
    
    # Create status window
    status_root = tk.Tk()
    status_root.title("Program Status")
    status_root.geometry("1200x1150")

    # Progress bar for set_pts
    ttk.Label(status_root, 
              text=f"Program Status: {programme.cal_type} {programme.clean_program.capitalize()}", 
              font=("Courier", 18), 
              justify="center").pack(pady=5)
    step_progress = ttk.Progressbar(status_root, maximum=len(set_pts[0]), length=500)
    step_progress.pack(pady=5)

    # Step progress label
    step_label = ttk.Label(status_root, text='Starter...', font=("Courier", 12), justify="left")
    step_label.pack(pady=5)

    # Progress bar for time within each step
    ttk.Label(status_root, text="Status for Nuværende Trin", font=("Courier", 14), justify="center").pack(pady=5)
    time_progress = ttk.Progressbar(status_root, maximum=step_time, length=500)
    time_progress.pack(pady=5)

    # Status label
    status_label = ttk.Label(status_root, text="Starter...", font=("Courier", 12), justify="center")
    status_label.pack(pady=10)

    # Abort flag
    abort_flag = tk.BooleanVar(value=False)
    ttk.Button(status_root, text="Afbryd", command=lambda: abort_flag.set(True)).pack(pady=5)
    
    # Create the matplotlib figure with two y-axes
    fig, ax1 = plt.subplots(figsize=(12, 8))
    ax2 = ax1.twinx()

    # Initial empty plot
    color1, color2 = 'g', 'b'
    line1, = ax1.plot([], [], f'{color1}-')
    line2, = ax2.plot([], [], f'{color2}-')

    bbox1 = ax1.yaxis.label.get_window_extent()
    bbox2 = ax2.yaxis.label.get_window_extent()
    center_y = (bbox1.y0 + bbox1.y1 + bbox2.y0 + bbox2.y1) / 4
    center_y_fig = center_y / fig.bbox.height

    ax1.set_ylabel('Span\n[mln/min]', color=color1)
    ax1.tick_params(axis='y', colors=color1)
    ax1.yaxis.label.set_rotation(0)
    ax1.yaxis.set_label_coords(-0.1, center_y_fig)
    ax1.spines['left'].set_color(color1)
    

    ax2.set_ylabel('Fortynding\n[ln/min]', color=color2)
    ax2.tick_params(axis='y', colors=color2)
    ax2.yaxis.label.set_rotation(0)
    ax2.yaxis.set_label_coords(1.1, center_y_fig*1.1)
    ax2.spines['right'].set_color(color2)
    ax2.spines['left'].set_visible(False)

    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m %H:%M:%S'))
    fig.autofmt_xdate()
    fig.tight_layout()

    # Embed the plot in the tkinter window
    canvas = FigureCanvasTkAgg(fig, master=status_root)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10)

    status_root.update()

    # Wait until start_time is reached
    while datetime.datetime.now() < start_time:
        if abort_flag.get():
            status_label.config(text="Programmet blev afbrudt før start.")
            status_root.update()
            time.sleep(2)
            status_root.destroy()

            end_setpoint_frac = end_setpoint_pct / 100
            bronkhorst_large.write_bronkhorst(206, bronkhorst_large.max_flow*end_setpoint_frac)
            bronkhorst_small.write_bronkhorst(206, bronkhorst_small.max_flow*end_setpoint_frac)
            return  # Exit the function cleanl

        remaining = start_time - datetime.datetime.now()
        hours, remainder = divmod(int(remaining.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)
        status_label.config(text=(
            f"Venter på starttidspunkt...\n"
            f"Starter om {hours:02d}:{minutes:02d}:{seconds:02d}\n"
            f"Valgte starttidspunkt er: {programme.selected_starttime.strftime("%m-%d %H:%M")}"
        ))
        status_root.update()
        time.sleep(1)


    # 206 is the DDE number for setting the specific flow of a Bronkhorst MFC
    for i, (dilution, span, conc) in enumerate(zip(*set_pts)):
        if not flow_list_large and not flow_list_small:
            flow_large = 0.00
            flow_small = 0.00
        else:
            flow_large = (flow_list_large[-1]/bronkhorst_large.max_flow)*100
            flow_small = (flow_list_small[-1]/bronkhorst_small.max_flow)*100
        dilution_flow = pct_mln_conversion(bronkhorst_large.max_flow, dilution)
        span_flow = pct_mln_conversion(bronkhorst_small.max_flow, span)
        bronkhorst_large.write_bronkhorst(206, dilution_flow)
        bronkhorst_small.write_bronkhorst(206, span_flow)

        # Update step progress
        tot_time_left = datetime.timedelta(seconds=(len(set_pts[0])-i*step_time))
        tot_hours, tot_remainder = divmod(int(tot_time_left.total_seconds()), 3600)
        tot_minutes, tot_seconds = divmod(tot_remainder, 60)

        step_progress['value'] = i + 1
        step_label.config(text=f"Trin {i+1}/{len(set_pts[0])}\tTid tilbage total: {tot_hours:02d}:{tot_minutes:02d}:{tot_seconds:02d}\n"
                               f"\n{'':<25}{'Indstillet':<15}{'Målt':<10}\n"
                               f"{'Fortynding:':<25}{f'{dilution}%':<15}{f'{flow_large}%':<10}\n"
                               f"{'Span:':<25}{f'{span}%':<15}{f'{flow_small}%':<10}\n"
                               f"{'Koncentration:':<25}{conc:.2f} ppb")
        status_root.update()

        for t in range(step_time):
            if abort_flag.get():
                status_label.config(text="Programmet blev afbrudt under kørsel.")
                status_root.update()

                # Save csv file
                csv_rows = zip(time_list, flow_list_small, flow_list_large)
                with open(f'Flow_logs/{programme.save_name}.csv', 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(csv_header)
                    writer.writerows(csv_rows)

                # Save plot
                fig.savefig(f'Figures/{programme.save_name}.pdf')

                time.sleep(2)
                status_root.destroy()
                end_setpoint_frac = end_setpoint_pct / 100
                bronkhorst_large.write_bronkhorst(206, bronkhorst_large.max_flow*end_setpoint_frac)
                bronkhorst_small.write_bronkhorst(206, bronkhorst_small.max_flow*end_setpoint_frac)
                return
            time_list.append(datetime.datetime.now())
            flow_list_small.append(read_bh_flow(bronkhorst_small))
            flow_list_large.append(read_bh_flow(bronkhorst_large)/1000)
            flow_large = (flow_list_large[-1]/bronkhorst_large.max_flow)*100
            flow_small = (flow_list_small[-1]/bronkhorst_small.max_flow)*100

            # Update plot data
            line1.set_data(time_list, flow_list_small)
            line2.set_data(time_list, flow_list_large)
            ax1.relim()
            ax1.autoscale_view()
            ax2.relim()
            ax2.autoscale_view()
            canvas.draw()

            tot_time_left = datetime.timedelta(seconds=(len(set_pts[0])-(i+1))*step_time+step_time-(t+1))
            tot_hours, tot_remainder = divmod(int(tot_time_left.total_seconds()), 3600)
            tot_minutes, tot_seconds = divmod(tot_remainder, 60)

            step_time_left = datetime.timedelta(seconds=step_time-(t+1))
            step_hours, step_remainder = divmod(int(step_time_left.total_seconds()), 3600)
            step_minutes, step_seconds = divmod(step_remainder, 60)

            time_progress['value'] = t + 1
            step_label.config(text=f"Trin {i+1}/{len(set_pts[0])}\tTid tilbage total: {tot_hours:02d}:{tot_minutes:02d}:{tot_seconds:02d}\n"
                                   f"\n{'':<25}{'Indstillet':<15}{'Målt':<10}\n"
                                   f"{'Fortynding:':<25}{f'{dilution}%':<15}{f'{flow_large:.2f}%':<10}\n"
                                   f"{'Span:':<25}{f'{span}%':<15}{f'{flow_small:.2f}%':<10}\n"
                                   f"{'Koncentration:':<25}{conc:.2f} ppb")
            status_label.config(text=f"Tid tilbage på trin: {step_hours:02d}:{step_minutes:02d}:{step_seconds:02d}")
            status_root.update()
            time.sleep(1)

        time_progress['value'] = 0  # Reset time progress for next step
    status_label.config(text="Program Færdig.\nGemmer nu Figur og csv fil.")
    status_root.update()

    # Save csv file
    csv_rows = zip(time_list, flow_list_small, flow_list_large)
    with open(f'Flow_logs/{programme.save_name}.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(csv_header)
        writer.writerows(csv_rows)
    
    # Save plot
    fig.savefig(f'Figures/{programme.save_name}.pdf')
    
    time.sleep(2)
    status_root.destroy()

    end_setpoint_frac = end_setpoint_pct / 100
    bronkhorst_large.write_bronkhorst(206, bronkhorst_large.max_flow*end_setpoint_frac)
    bronkhorst_small.write_bronkhorst(206, bronkhorst_small.max_flow*end_setpoint_frac)
    print('Finished all')


def find_setpoints(programme: ProgrammeSelector):
    df = pd.read_excel(programme.ark)
    indexes = find_program_index(programme.ark, programme.program)
    setpoint1_name, setpoint2_name, ppb_names = programme.set_point_names
    setp1 = df[setpoint1_name].iloc[indexes].values
    setp2 = df[setpoint2_name].iloc[indexes].values
    conc_vals = df[ppb_names].iloc[indexes].values
    return setp1, setp2, conc_vals


if __name__ == '__main__':
    end_setpoints_pct = 1 # % of max flow

    # Find and connect the Bronkhorst MFC's
    bh_ports = list(find_bronkhorst_ports().values())
    bronkhorsts = [BronkhorstMFC(bh_port) for bh_port in bh_ports]

    # Find and load programme variables
    programme_variables = ProgrammeSelector()
    main_controller(bronkhorsts, programme_variables, end_setpoints_pct)