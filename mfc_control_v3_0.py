import os
import sys
import csv
import glob
import time
import random
import openpyxl
import datetime
import numpy as np
import pandas as pd
import tkinter as tk
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from tkinter import ttk
from tkinter import messagebox
from tkcalendar import Calendar
from bronkhorst_mfc_test.airpy import *
from bronkhorst_mfc_test.mfc_logger_v1 import *
from bronkhorst_mfc_test.mfc_controller_v1 import *
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter.filedialog import askopenfilename
from tkinter.scrolledtext import ScrolledText

version = 3.0

class ProgrammeSelector:
    def __init__(self):
        cal_types = {'SO2': ['Fd %', 'ml/min', 'Fso2', 'ml/min ', 'SO2 ppb'],
                     'NOx': ['Fd %', 'Fno', 'No PPB', 'Nox PPB']}
        # File selection window
        root_file = tk.Tk()
        root_file.withdraw()
        root_file.wm_attributes('-topmost', 1)
        self.ark = askopenfilename(parent=root_file, 
                                   title='Vælg fil med instrukserne',
                                   filetypes=[('Excel files', '*.xlsx')])
        root_file.destroy()

        # Load and slice DataFrame
        try:
            df = pd.read_excel(self.ark)
            for idx, item in enumerate(df.iloc[:, 0]):
                try:
                    int(item)
                except ValueError:
                    slicer = idx
                    break
            df = df.loc[:slicer-1]
        except Exception as e:
            print(f'Error reading Excel file: {e}')
            return
        self.options = list(df.columns[-3:])
        self.options.append('Nul')

        # Sequential GUI flow
        self.program, self.cal_type, self.clean_program, self.shuffle_steps = self._select_programme()
        self.program_steps = self._find_program_index()
        self.selected_starttime, self.time_per_step = self._select_time()
        self.set_point_names = cal_types[self.cal_type]
        self.save_name = f'{self.selected_starttime.strftime("%Y_%m_%d_%H_%M")}_{self.cal_type}_{self.clean_program}'
    

    def _select_programme(self):
        selected_programme = []
        selected_species = []
        selected_shuffle = []

        def _confirm_selection():
            selected_prog = programme_var.get()
            selected_spec = species_var.get()
            selected_shuf = shuffle_var.get()
            selected_programme.append(selected_prog)
            selected_species.append(selected_spec)
            selected_shuffle.append(selected_shuf)
            if selected_prog and selected_spec:
                root_select.destroy()
            else:
                messagebox.showwarning('Ugyldigt valg!', 'Vælg venligst både et program og en type og tryk bekræft.')

        root_select = tk.Tk()
        root_select.wm_attributes('-topmost', 1)
        root_select.title('Vælg Program og Type')
        root_select.geometry('500x300')

        # Main frame with two columns
        main_frame = ttk.Frame(root_select)
        main_frame.pack(expand=True, fill='both', padx=10, pady=10)

        shuffle_var = tk.BooleanVar(value=False)  
        shuffle_frame = ttk.LabelFrame(main_frame, text='Rækkefølge')
        shuffle_frame.pack(side='top', fill='x', pady=(0, 10))

        ttk.Radiobutton(shuffle_frame, text='Tilfældig', variable=shuffle_var, value=True).pack(side='left', padx=10, pady=5)
        ttk.Radiobutton(shuffle_frame, text='Ikke Tilfældig', variable=shuffle_var, value=False).pack(side='left', padx=10, pady=5)

        # Species selection
        species_var = tk.StringVar(value='SO2')
        species_frame = ttk.LabelFrame(main_frame, text='Vælg Type')
        species_frame.pack(side='left', expand=True, fill='both', padx=10)
        ttk.Radiobutton(species_frame, text='SO₂', variable=species_var, value='SO2').pack(anchor='w', pady=5)
        ttk.Radiobutton(species_frame, text='NOx', variable=species_var, value='NOx').pack(anchor='w', pady=5)

        # Programme selection
        programme_var = tk.StringVar(value='Nul')
        programme_frame = ttk.LabelFrame(main_frame, text='Vælg Program')
        programme_frame.pack(side='right', expand=True, fill='both', padx=10)
        
        clean_text_dict = {}
        for option in self.options:
            if 'skift' in option.lower():
                option_text = 'Reference Gas Skift'
            else:
                option_text = option
            clean_text_dict[option] = option_text
            ttk.Radiobutton(programme_frame, text=option_text, variable=programme_var, value=option).pack(anchor='w', pady=5)

        # Confirm button
        ttk.Button(root_select, text='Bekræft', command=_confirm_selection).pack(pady=10)

        root_select.mainloop()
        return selected_programme[0], selected_species[0], clean_text_dict[selected_programme[0]].replace(' ', '_').lower(), selected_shuffle[0]


    def _find_program_index(self):
        if 'Nul' in self.program:
            return None
        
        workbook = openpyxl.load_workbook(filename=self.ark, data_only=True)
        sheet = workbook.worksheets[0]
        data = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx < 1:
                header = row
            elif 1 <= idx <= 51:
                data.append(row)

        df = pd.DataFrame(data, columns=header)
        for row in sheet.iter_rows(max_row=51):
            for cell in row:
                cell_color = cell.fill.start_color.index
                if str(cell_color) not in ['00000000', '0', '43']:
                    df.iat[cell.row-2, cell.column-1] = 'Colored'
                else:
                    df.iat[cell.row-2, cell.column-1] = None

        if 'Skift' not in self.program:
            return df[df[self.program] == 'Colored'].index
        else:
            return df[df['Inds v skift af ref'] == 'Colored'].index


    def _select_time(self):
        selected_datetime = []

        def _confirm_selection():
            if start_mode.get() == 'later':
                selected_date = cal.selection_get()
                selected_time = datetime.time(hour=hour_var.get(), minute=minute_var.get())
                selected_datetime.append(datetime.datetime.combine(selected_date, selected_time))
            else:
                selected_datetime.append(datetime.datetime.now())
            #print(f'Selected Start Time: {selected_datetime}')
            root_time.destroy()

        def _update_ui():
            is_later = start_mode.get() == 'later'
            cal.configure(state='normal' if is_later else 'disabled')
            hour_spinbox.configure(state='normal' if is_later else 'disabled')
            minute_spinbox.configure(state='normal' if is_later else 'disabled')

        root_time = tk.Tk()
        root_time.title('Vælg starttidspunkt for programmet')
        root_time.geometry('450x600')

        # Start mode selection
        start_mode = tk.StringVar(value='now')
        mode_frame = ttk.LabelFrame(root_time, text='Vælg Start')
        mode_frame.pack(padx=10, pady=10, fill='x')

        ttk.Radiobutton(mode_frame, text='Start Nu', variable=start_mode, value='now', command=_update_ui).pack(anchor='w', padx=10)
        ttk.Radiobutton(mode_frame, text='Start Senere', variable=start_mode, value='later', command=_update_ui).pack(anchor='w', padx=10)

        # Calendar
        cal_frame = ttk.LabelFrame(root_time, text='Vælg Dato')
        cal_frame.pack(padx=10, pady=10, fill='both', expand=True)
        cal = Calendar(cal_frame, selectmode='day', date_pattern='yyyy-mm-dd')
        cal.pack(padx=10, pady=10)

        # Time selection
        hour_var = tk.IntVar(value=datetime.datetime.now().hour)
        minute_var = tk.IntVar(value=datetime.datetime.now().minute)

        time_frame = ttk.LabelFrame(root_time, text='Vælg Tid')
        time_frame.pack(padx=10, pady=10, fill='x')
        time_inner_frame = ttk.Frame(time_frame)
        time_inner_frame.pack(pady=10)

        hour_spinbox = ttk.Spinbox(time_inner_frame, from_=0, to=23, textvariable=hour_var, width=5, justify='center')
        hour_spinbox.pack(side='left')
        ttk.Label(time_inner_frame, text=':', font=('Segoe UI', 12)).pack(side='left', padx=5)
        minute_spinbox = ttk.Spinbox(time_inner_frame, from_=0, to=59, textvariable=minute_var, width=5, justify='center')
        minute_spinbox.pack(side='left')

        # Time at each step
        step_time_var = tk.IntVar(value=50)
        step_frame = ttk.LabelFrame(root_time, text='Tid pr. trin (minuter)')
        step_frame.pack(padx=10, pady=10, fill='x')
        ttk.Spinbox(step_frame, from_=1, to=120, textvariable=step_time_var, width=5).pack(pady=5)

        # Expected finish time
        expected_label = ttk.Label(root_time, text='Forventet færdig: --:--')
        expected_label.pack(pady=10)

        def _update_expected_time(*args):
            if start_mode.get() == 'later':
                try:
                    start_dt = datetime.datetime.combine(cal.selection_get(), datetime.time(hour=hour_var.get(), minute=minute_var.get()))
                except Exception:
                    return
            else:
                start_dt = datetime.datetime.now()

            steps = getattr(self, 'program_steps', [])
            if steps is not None:
                n_steps = len(steps) + 2 # Accounts for the Zero befor and after measuring
            else:
                n_steps = 1
            try:
                step_time_val = step_time_var.get()
                if isinstance(step_time_val, int):
                    duration = datetime.timedelta(minutes=step_time_val * n_steps)
                    expected_dt = start_dt + duration
                    expected_label.config(text=f'Forventet færdig: {expected_dt.strftime('%Y-%m-%d %H:%M')}')
                else:
                    expected_label.config(text='Forventet færdig: N/A')
            except tk.TclError:
                expected_label.config(text='Forventet færdig: N/A')


        # Bind updates
        step_time_var.trace_add('write', _update_expected_time)
        hour_var.trace_add('write', _update_expected_time)
        minute_var.trace_add('write', _update_expected_time)
        start_mode.trace_add('write', _update_expected_time)
        cal.bind('<<CalendarSelected>>', lambda e: _update_expected_time())

        # Confirm button
        ttk.Button(root_time, text='Bekræft', command=_confirm_selection).pack(pady=10)

        _update_ui()
        _update_expected_time()

        root_time.mainloop()
        time_per_step = step_time_var.get()
        return selected_datetime[0], time_per_step


def start_waiter(start_time,
                  abort_flag, 
                  status_root, 
                  status_label, 
                  programme, 
                  end_setpoint_frac,
                  bronkhorst_small,
                  bronkhorst_large):
    
    # Wait until start_time is reached
    bh_small_idle_point = bronkhorst_small.max_flow*end_setpoint_frac[0]
    bh_large_idle_point = bronkhorst_large.max_flow*end_setpoint_frac[1]

    while datetime.datetime.now() < start_time:
        if abort_flag.get():
            status_label.config(text='Programmet blev afbrudt før start.')
            status_root.update()
            time.sleep(2)
            status_root.destroy()
            bronkhorst_small.write_bronkhorst(206, bh_small_idle_point)
            bronkhorst_large.write_bronkhorst(206, bh_large_idle_point)
            return  # Exit the function cleanl

        remaining = start_time - datetime.datetime.now()
        hours, remainder = divmod(int(remaining.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)
        status_label.config(text=(
            f'Venter på starttidspunkt...\n'
            f'Starter om {hours:02d}:{minutes:02d}:{seconds:02d}\n'
            f'Valgte starttidspunkt er: {programme.selected_starttime.strftime('%d-%m %H:%M')}'
        ))
        status_root.update()
        time.sleep(1)


def initiate_plot(plot_frame, 
                  bronkhorst_small, 
                  bronkhorst_large):
    # Create the matplotlib figure with two y-axes
    fig, ax1 = plt.subplots(figsize=(10, 7.5))
    ax2 = ax1.twinx()

    # Initial empty plot
    color1, color2 = 'g', 'b'
    line1, = ax1.plot([], [], f'{color1}-')
    line2, = ax2.plot([], [], f'{color2}-')

    bbox1 = ax1.yaxis.label.get_window_extent()
    bbox2 = ax2.yaxis.label.get_window_extent()
    center_y = (bbox1.y0 + bbox1.y1 + bbox2.y0 + bbox2.y1) / 4
    center_y_fig = center_y / fig.bbox.height

    ax1.set_ylabel(f'Span\n[{bronkhorst_small.pretty_unit}]', color=color1)
    ax1.tick_params(axis='y', colors=color1)
    ax1.yaxis.label.set_rotation(0)
    ax1.yaxis.set_label_coords(-0.1, center_y_fig)
    ax1.spines['left'].set_color(color1)
    

    ax2.set_ylabel(f'Fortynding\n[{bronkhorst_large.pretty_unit}]', color=color2)
    ax2.tick_params(axis='y', colors=color2)
    ax2.yaxis.label.set_rotation(0)
    ax2.yaxis.set_label_coords(1.1, center_y_fig*1.1)
    ax2.spines['right'].set_color(color2)
    ax2.spines['left'].set_visible(False)

    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m %H:%M:%S'))
    fig.autofmt_xdate()
    fig.tight_layout()

    # Plot goes in plot_frame
    canvas = FigureCanvasTkAgg(fig, master=plot_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)

    return fig, ax1, ax2, line1, line2, canvas


def on_programme_complete(info_frame, abort_button, status_root):
    # Destroys and replaces the abort button with a Finish button
    # After program completion
    abort_button.destroy()

    done_button = ttk.Button(info_frame, text='Afslut', command=lambda: (status_root.destroy, sys.exit()))
    done_button.pack(pady=5)


def load_file_contents(text_display, comment_file_name):
    try:
        with open(comment_file_name, 'r', encoding='utf-8') as f:
            contents = f.read()
            text_display.configure(state='normal')
            text_display.delete('1.0', tk.END)
            text_display.insert(tk.END, contents)
            text_display.configure(state='disabled')
    except Exception as e:
        print(f"Error loading file: {e}")


def initiate_window(programme, step_time, final_point_list, shuffled, comment_file_name):
    # Create main window
    status_root = tk.Tk()
    status_root.title(f'MFC Control {version}')
    status_root.geometry('1200x1150')

    # Create a top-level container frame to hold info_frame and input_frame side-by-side
    top_frame = ttk.Frame(status_root)
    top_frame.pack(side='top', fill='x', pady=10)

    # Create info_frame on the left
    info_frame = ttk.Frame(top_frame)
    info_frame.pack(side='left', fill='y', padx=50)

    # Create input_frame on the right
    input_frame = ttk.Frame(top_frame)
    input_frame.pack(side='right', padx=10, pady=10, anchor='ne', expand=True, fill='y')

    # Plot frame remains below
    plot_frame = ttk.Frame(status_root)
    plot_frame.pack(side='top', fill='both', expand=True)

    # Add a spacer to push content to the bottom
    ttk.Label(input_frame, 
              text='Indhold i Kommentarfil', 
              font=('Courier', 14), 
              justify='center').pack(side='top', fill='both', expand=True)

    # Text display field (read-only)
    text_display = ScrolledText(input_frame, height=15, width=60, font=('Courier', 10))
    text_display.pack(side='top', pady=(0, 10))
    text_display.configure(state='disabled')  # Make it read-only
    load_file_contents(text_display, comment_file_name)

    # Create bottom-aligned inner frame
    bottom_input_frame = ttk.Frame(input_frame)
    bottom_input_frame.pack(side='bottom', pady=10, fill='x')

    # Add a spacer to push content to the bottom
    ttk.Label(bottom_input_frame, 
              text='Skriv Kommentarer', 
              font=('Courier', 12), 
              justify='center').pack(side='top', pady=(0, 5))

    # Entry widget for text input
    input_entry = ttk.Entry(bottom_input_frame, width=80)
    input_entry.pack(side='top', pady=(5, 0))

    # Progress bar for set_pts
    ttk.Label(info_frame, 
              text=f'Program: {programme.cal_type} {" ".join([i.capitalize() for i in programme.clean_program.split("_")])} {shuffled}', 
              font=('Courier', 18), 
              justify='center').pack(pady=5)

    step_progress = ttk.Progressbar(info_frame, maximum=len(final_point_list), length=600)
    step_progress.pack(pady=5)

    # Step progress label
    step_label = ttk.Label(info_frame, text='Starter...', font=('Courier', 12), justify='left')
    step_label.pack(pady=5)

    # Progress bar for time within each step
    ttk.Label(info_frame, text='Status for Nuværende Trin', font=('Courier', 14), justify='center').pack(pady=5)
    time_progress = ttk.Progressbar(info_frame, maximum=step_time, length=600)
    time_progress.pack(pady=5)

    # Status label
    status_label = ttk.Label(info_frame, text='Starter...', font=('Courier', 12), justify='center')
    status_label.pack(pady=10)

    # Abort flag
    abort_flag = tk.BooleanVar(value=False)
    abort_button = ttk.Button(info_frame, text='Afbryd', command=lambda: abort_flag.set(True))
    abort_button.pack(pady=5)
    status_root.update()

    return (status_root, 
            info_frame, 
            plot_frame, 
            bottom_input_frame,
            input_entry,
            step_progress, 
            step_label, 
            time_progress, 
            status_label, 
            abort_flag, 
            abort_button,
            text_display)


def flow_controller(bronkhorsts: list[BronkhorstMFC], 
                    programme: ProgrammeSelector, 
                    end_setpoint_frac: int) -> None:
    '''
    Defines the main function to controll the Bronkhorst MFC's using the worksheet.

    :param bronkhorsts: List of Bronkhorst MFC objects to be controlled.
    :param sleep_time: Time the program is supposed to sleep between dilution steps
                       to achieve a stable concentration.
    '''

    if len(bronkhorsts) != 2:
        raise KeyError('Uncompatible number of Bronkhorst MFCs connected. Program can only handle 2 MFCs (span + dilution).')
    
    def normalize_flow(mfc: BronkhorstMFC):
        # Convert everything to ln/min for comparison
        if mfc.readout_unit == 'mln/min':
            return mfc.max_flow / 1000
        return mfc.max_flow
    
    # Starting time and flows
    start_time = programme.selected_starttime
    time_list = []
    flow_list_small = []
    flow_list_large = []
    
    # Ensure folders exist
    os.makedirs(f'{programme.save_name}', exist_ok=True)

    # Log/comment file
    comment_file_name = f'{programme.save_name}/comments_description.txt'
    setting_text = ['{:<16}{:<16}{:<20}{:<12}{:<18}{:<20}'.format('Tidspunkt', 
                                                                  'Fortynding [%]', 
                                                                  'Fortynding [L/min]', 
                                                                  'Span [%]', 
                                                                  'Span [mL/min]', 
                                                                  'Koncentration [ppb]')]
    
    # Initialize the file (create if it doesn't exist)
    with open(comment_file_name, 'a') as f:
        f.write(f'Kommentarer og information til {programme.cal_type} '
                f'{' '.join([i.capitalize() for i in programme.clean_program.split('_')])}\n')

    # Function to append text to the file
    def append_to_file(text):
        with open(comment_file_name, 'a') as f:
            f.write(text + '\n')

    # Sort the list by normalized flow
    mfcs_sorted = sorted(bronkhorsts, key=normalize_flow)
    bronkhorst_small = mfcs_sorted[0]
    bronkhorst_large = mfcs_sorted[1]
    bh_small_idle_point = bronkhorst_small.max_flow*end_setpoint_frac[0]
    bh_large_idle_point = bronkhorst_large.max_flow*end_setpoint_frac[1]
    set_large, flow_large, set_small, flow_small, ppb_conc = find_setpoints(programme)

    set_pts = (set_large, flow_large, set_small, flow_small, ppb_conc)
    set_pt_list = [i for i in zip(*set_pts)]

    if len(set_pt_list) > 1 and programme.shuffle_steps:
        shuffled = 'Tilfældig'
        final_point_list = random.sample(set_pt_list, len(set_pt_list))
        final_point_list.insert(0, (90, np.max(flow_large), 0, 0, 0))
        final_point_list.append((90, np.max(flow_large), 0, 0, 0))

    elif len(set_pt_list) > 1 and not programme.shuffle_steps:
        shuffled = 'Ikke Tilfældig'
        final_point_list = set_pt_list
        final_point_list.insert(0, (90, np.max(flow_large), 0, 0, 0))
        final_point_list.append((90, np.max(flow_large), 0, 0, 0))
    
    else:
        final_point_list = set_pt_list
        shuffled = ''

    step_time = programme.time_per_step*60
    csv_header = ['Datetime', 
                  f'Bronkhorst {bronkhorst_small.max_flow:.1f}SCCM [mL/min]', 
                  f'Bronkhorst {bronkhorst_large.max_flow:.1f}SLM [L/min]']
    
    (status_root,     # Main window
     info_frame,      # Frame for UI elements
     plot_frame,      # Frame for matplotlib plot,
     input_frame,     # Frame for comments
     input_entry,     # Input in the field
     step_progress,   # Progress bar for steps
     step_label,      # Label showing current step
     time_progress,   # Progress bar for time
     status_label,    # Status message
     abort_flag,      # BooleanVar for abort
     abort_button,    # Abort button
     text_display     # Comment file text window
    ) = initiate_window(programme, step_time, final_point_list, shuffled, comment_file_name)
    
    (fig,       # Figure
     ax1,       # Axis for small MFC
     ax2,       # Axis for large MFC
     line1,     # Line for small MFC
     line2,     # Line for large MFC
     canvas     # Tkinter canvas for plot
    ) = initiate_plot(plot_frame, bronkhorst_small, bronkhorst_large)

    # Submit button
    def on_submit():
        text = f'{datetime.datetime.now().strftime('%d/%m-%Y %H:%M:%S')} {input_entry.get()}'
        append_to_file(text)
        input_entry.delete(0, tk.END)
        load_file_contents(text_display, comment_file_name)
      # Clear the entry after submission
    
    submit_button = ttk.Button(input_frame, text='Submit', command=on_submit)
    submit_button.pack(side='top')

    start_waiter(start_time, abort_flag, status_root, status_label, 
                 programme, end_setpoint_frac, bronkhorst_small, bronkhorst_large)

    # 206 is the DDE number for setting the specific flow of a Bronkhorst MFC
    for i, (dilution, dilution_flow, span, span_flow, conc) in enumerate(final_point_list):
        if not flow_list_large and not flow_list_small:
            flow_large = read_bh_flow(bronkhorst_small)
            flow_small = read_bh_flow(bronkhorst_large)/1000
        else:
            flow_large = (flow_list_large[-1]/bronkhorst_large.max_flow)*100
            flow_small = (flow_list_small[-1]/bronkhorst_small.max_flow)*100
        dilution_flow = pct_mln_conversion(bronkhorst_large.max_flow, dilution)
        span_flow = pct_mln_conversion(bronkhorst_small.max_flow, span)
        bronkhorst_large.write_bronkhorst(206, dilution_flow)
        bronkhorst_small.write_bronkhorst(206, span_flow)

        # Append new settings to comment/log file
        setting_text.append('{:<16}{:<16}{:<20.2f}{:<12}{:<18.2f}{:<20.2f}'.format(datetime.datetime.now().strftime('%d/%m %H:%M'), dilution, dilution_flow, span, span_flow, conc))
        
        # Update step progress
        tot_time_left = datetime.timedelta(seconds=(len(final_point_list)-i*step_time))
        tot_hours, tot_remainder = divmod(int(tot_time_left.total_seconds()), 3600)
        tot_minutes, tot_seconds = divmod(tot_remainder, 60)

        step_progress['value'] = i + 1
        step_label.config(text=f'Trin{i+1}/{len(final_point_list)}\t  Tid tilbage total: {tot_hours:02d}:{tot_minutes:02d}:{tot_seconds:02d}\n'
                               f'Forventet færdig: {(datetime.datetime.now()+tot_time_left).strftime("%d-%m %H:%M")}\n'
                               f'\n{"":<17}{"Indstillet":<7}{"":<15}{"Målt":<10}\n'
                               f'{"Fortynding:":<15}{f"{dilution}%":<5}{f"{dilution_flow} L/min":<20}{f"{dilution}%":<5}{f"{dilution_flow} L/min":<10}\n'
                               f'{"Span:":<15}{f"{span}%":<5}{f"{span_flow} mL/min":<20}{f"{span}%":<5}{f"{span_flow} mL/min":<10}\n'
                               f'{"Koncentration:":<15}{conc:.2f} ppb')
        status_root.update()

        for t in range(step_time):
            if abort_flag.get():
                status_label.config(text='Programmet blev afbrudt under kørsel.')
                status_root.update()

                # Ensure folders exist
                os.makedirs(f'{programme.save_name}', exist_ok=True)

                # Save csv file
                csv_rows = zip(time_list, flow_list_small, flow_list_large)
                with open(f'Flow_logs/{programme.save_name}.csv', 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(csv_header)
                    writer.writerows(csv_rows)

                # Save plot
                fig.savefig(f'Figures/{programme.save_name}.pdf')

                time.sleep(2)
                status_root.destroy()
                bronkhorst_large.write_bronkhorst(206, bh_small_idle_point)
                bronkhorst_small.write_bronkhorst(206, bh_large_idle_point)
                return
            
            time_list.append(datetime.datetime.now())
            flow_list_small.append(read_bh_flow(bronkhorst_small))
            flow_list_large.append(read_bh_flow(bronkhorst_large)/1000)
            flow_large = (flow_list_large[-1]/bronkhorst_large.max_flow)*100
            flow_small = (flow_list_small[-1]/bronkhorst_small.max_flow)*100

            # Update plot data
            line1.set_data(time_list, flow_list_small)
            line2.set_data(time_list, flow_list_large)
            ax1.relim()
            ax1.set_ylim([0, np.float16(f'{bronkhorst_large.max_flow:.2f}')])
            ax1.autoscale_view()
            ax2.relim()
            ax2.set_ylim([0, np.float16(f'{bronkhorst_large.max_flow:.2f}')])
            ax2.autoscale_view()
            canvas.draw()

            tot_time_left = datetime.timedelta(seconds=(len(final_point_list)-(i+1))*step_time+step_time-(t+1))
            tot_hours, tot_remainder = divmod(int(tot_time_left.total_seconds()), 3600)
            tot_minutes, tot_seconds = divmod(tot_remainder, 60)

            step_time_left = datetime.timedelta(seconds=step_time-(t+1))
            step_hours, step_remainder = divmod(int(step_time_left.total_seconds()), 3600)
            step_minutes, step_seconds = divmod(step_remainder, 60)

            time_progress['value'] = t + 1
            step_label.config(text=f"Trin {i+1}/{len(final_point_list)}\t  Tid tilbage total: {tot_hours:02d}:{tot_minutes:02d}:{tot_seconds:02d}\n"
                               f"Forventet færdig: {(datetime.datetime.now()+tot_time_left).strftime("%d-%m %H:%M")}\n"
                               f"\n{'':<17}{'Indstillet':<7}{'':<15}{'Målt':<10}\n"
                               f"{'Fortynding:':<15}{f'{dilution}%':<5}{f'{dilution_flow} L/min':<20}{f'{dilution}%':<5}{f'{dilution_flow} L/min':<10}\n"
                               f"{'Span:':<15}{f'{span}%':<5}{f'{span_flow} mL/min':<20}{f'{span}%':<5}{f'{span_flow} mL/min':<10}\n"
                               f"{'Koncentration:':<15}{conc:.2f} ppb")
            status_label.config(text=f'Tid tilbage på trin: {step_hours:02d}:{step_minutes:02d}:{step_seconds:02d}')
            status_root.update()
            time.sleep(1)

        time_progress['value'] = 0  # Reset time progress for next step
        
    status_label.config(text='Program Færdig.\nGemmer nu Figur og csv fil.')
    status_root.update()

    # Ensure folders exist
    os.makedirs(f'{programme.save_name}', exist_ok=True)
    
    # Save csv file
    csv_rows = zip(time_list, flow_list_small, flow_list_large)
    with open(f'{programme.save_name}/flow_plot{programme.selected_starttime.strftime("%d_%m_%H_%M")}.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(csv_header)
        writer.writerows(csv_rows)

    # Save plot
    fig.savefig(f'{programme.save_name}/flow_plot_{programme.selected_starttime.strftime("%d_%m_%H_%M")}.pdf')

    time.sleep(2)
    comment_settings = '\n'.join(setting_text)
    append_to_file(comment_settings)
    on_programme_complete(info_frame, abort_button, status_root)

    bronkhorst_small.write_bronkhorst(206, bh_small_idle_point)
    bronkhorst_large.write_bronkhorst(206, bh_large_idle_point)
    print('Finished all')
    status_root.mainloop()


def find_setpoints(programme: ProgrammeSelector):
    df = pd.read_excel(programme.ark)
    setpoint1_name, setflow1_name, setpoint2_name, setflow2_name, ppb_names = programme.set_point_names
    if programme.program == 'Nul':
        set_large = [90]
        set_small = [0]
        conc_vals = [0]
        flow_small = [0]
        flow_large = [df[setflow1_name][df[setpoint1_name]==90].iloc[0]]
        return set_large, flow_large, set_small, flow_small, conc_vals
    else:
        indexes = find_program_index(programme.ark, programme.program)
        set_large = df[setpoint1_name].iloc[indexes].values
        set_small = df[setpoint2_name].iloc[indexes].values
        flow_large = df[setflow1_name].iloc[indexes].values
        flow_small = df[setflow2_name].iloc[indexes].values
        conc_vals = df[ppb_names].iloc[indexes].values
        return set_large, flow_large, set_small, flow_small, conc_vals, 


if __name__ == '__main__':
    end_setpoints_frac = [0.01, 0.6] # % of max flow

    # Find and connect the Bronkhorst MFC's
    bh_ports = list(find_bronkhorst_ports().values())
    bronkhorsts = [BronkhorstMFC(bh_port) for bh_port in bh_ports]

    # Find and load programme variables
    programme_variables = ProgrammeSelector()
    flow_controller(bronkhorsts, programme_variables, end_setpoints_frac)